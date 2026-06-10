from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def test_marketplaces_have_dedicated_generators():
    ui = (ROOT / "frontend" / "app_streamlit.py").read_text(encoding="utf-8")

    assert "scripts/run_generator_myntra.py" in ui
    assert "scripts/run_generator_flipkart.py" in ui
    assert "Myntra_Sku_Ready" in ui
    assert "Flipkart_Sku_Ready" in ui
    assert 'OUTPUT_DIR / "flipkart"' in ui


def test_marketplace_templates_are_distinct():
    myntra_path = ROOT / "data" / "input" / "Myntra-Sku-Template-2026-02-21.xlsx"
    flipkart_path = ROOT / "data" / "input" / "Flipkart_Bra_Template.xlsx"

    assert myntra_path.exists()
    assert flipkart_path.exists()
    assert myntra_path.read_bytes() != flipkart_path.read_bytes()

    myntra = load_workbook(myntra_path, read_only=True, data_only=True)
    flipkart = load_workbook(flipkart_path, read_only=True, data_only=True)
    myntra_sheet = myntra["Bra"] if "Bra" in myntra.sheetnames else myntra.active
    flipkart_sheet = flipkart["Bra"] if "Bra" in flipkart.sheetnames else flipkart.active

    myntra_headers = {
        str(cell.value).strip()
        for cell in myntra_sheet[3]
        if cell.value is not None
    }
    flipkart_headers = {
        str(cell.value).strip()
        for cell in flipkart_sheet[1]
        if cell.value is not None
    }

    assert "vendorSkuCode" in myntra_headers
    assert "Seller SKU ID" in flipkart_headers
    assert myntra_headers != flipkart_headers

