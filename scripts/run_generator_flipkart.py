"""
FLIPKART Listing Generator
Dedicated script for Flipkart platform — separate from Myntra logic.
Template: header row 1, data starts row 5.
Brand keys use fk_ prefix (fk_komli, fk_invisisoft, etc.)
"""

import json
import yaml
import sys
import os
import socket
from pathlib import Path
import pickle
import time
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
import traceback

from engine.excel_reader import ExcelReader
from engine.rule_engine import RuleEngine
from engine.sku_generator import SkuGenerator
from engine.variant_generator import VariantGenerator
from engine.validator import Validator, ValidationError

# Multi-workbook support
try:
    from engine.multi_workbook_reader import MultiWorkbookReader, BrandConfigManager
    MULTI_WORKBOOK_AVAILABLE = True
except ImportError:
    MULTI_WORKBOOK_AVAILABLE = False
    MultiWorkbookReader = None
    BrandConfigManager = None

# Generation status reporting
try:
    from engine.generation_status import GenerationStatus, StatusLevel
    STATUS_REPORTING_AVAILABLE = True
except ImportError:
    STATUS_REPORTING_AVAILABLE = False
    GenerationStatus = None
    StatusLevel = None

try:
    from ai_layer.device_client import DeviceClient
    DEVICE_CLIENT_AVAILABLE = True
except ImportError:
    DeviceClient = None
    DEVICE_CLIENT_AVAILABLE = False

# Dynamic template loader
try:
    from engine.dynamic_template_loader import DynamicTemplateLoader
    DYNAMIC_TEMPLATE_AVAILABLE = True
except ImportError:
    DYNAMIC_TEMPLATE_AVAILABLE = False
    DynamicTemplateLoader = None

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy

# ============ CONFIG ============

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_DIR = BASE_DIR / "config"
INPUT_DIR = BASE_DIR / "data" / "input"
OUTPUT_DIR = BASE_DIR / "data" / "output"
CACHE_DIR = OUTPUT_DIR / ".cache"

PLATFORM = "FLIPKART"
ALL_VALID_BRANDS = [
    "fk_komli", "fk_invisisoft", "fk_tweens", "fk_dressberry",
    "fk_joomie", "fk_souminie", "fk_invisifit", "fk_intimist"
]

# Flipkart template rows
FLIPKART_HEADER_ROW = 1   # Headers are in Excel row 1
FLIPKART_START_ROW  = 5   # Data starts at Excel row 5

# Template state (set at runtime)
FLIPKART_TEMPLATE       = None
FLIPKART_TEMPLATE_SHEET = None

_template_loader = None


def get_template_loader():
    global _template_loader
    if _template_loader is None and DYNAMIC_TEMPLATE_AVAILABLE:
        _template_loader = DynamicTemplateLoader(str(INPUT_DIR))
    return _template_loader


def initialize_template():
    global FLIPKART_TEMPLATE, FLIPKART_TEMPLATE_SHEET
    fallback = INPUT_DIR / "Flipkart-Sku-Template.xlsx"
    loader   = get_template_loader()

    template_path = None
    if loader:
        try:
            template_path = loader.find_latest_flipkart_template()
        except Exception as e:
            print(f"  [WARNING] Template finder error: {e}")

    if template_path is None:
        print("  [WARNING] No Flipkart template found in data/input/, using fallback name")
        template_path = fallback

    sheet_name = None
    if loader:
        try:
            sheet_name = loader.get_template_sheet_name(template_path)
        except Exception:
            pass
    if sheet_name is None:
        sheet_name = "Bra"

    FLIPKART_TEMPLATE       = template_path
    FLIPKART_TEMPLATE_SHEET = sheet_name
    print(f"  [TEMPLATE] Flipkart -> {template_path.name} (sheet: {sheet_name})")


# ============ DEVICE CLIENT ============

def initialize_device_client():
    if not DEVICE_CLIENT_AVAILABLE:
        return None
    try:
        client = DeviceClient(auto_register=False)
        device_name = os.getenv("DEVICE_NAME") or socket.gethostname()
        if not client.is_registered():
            client.register_device(device_name=device_name, os_type=os.getenv("OS_TYPE", "windows"))
        return client
    except Exception as exc:
        print(f"[KILL SWITCH] Could not initialize device client: {exc}")
        return None


def ensure_not_terminated(client, stage):
    if not client:
        return
    try:
        if client.check_termination_status():
            print(f"[KILL SWITCH] Termination signal received during {stage}. Stopping.")
            sys.exit(1)
    except Exception as exc:
        print(f"[KILL SWITCH] Warning at {stage}: {exc}")


# ============ CACHING ============

ENABLE_CACHE = True
CACHE_DIR.mkdir(exist_ok=True, parents=True)


def get_cache_path(name):
    return CACHE_DIR / f"{name}.pkl"


def load_cached(name, loader_func, max_age_seconds=3600):
    if not ENABLE_CACHE:
        return loader_func()
    cache_path = get_cache_path(name)
    if cache_path.exists():
        age = time.time() - cache_path.stat().st_mtime
        if age < max_age_seconds:
            try:
                with open(cache_path, 'rb') as f:
                    print(f"  [OK] Loaded from cache: {name} (age: {age:.0f}s)")
                    return pickle.load(f)
            except Exception as e:
                print(f"  [WARN] Cache read failed: {e}, regenerating...")
    data = loader_func()
    try:
        with open(cache_path, 'wb') as f:
            pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
        print(f"  [OK] Saved to cache: {name}")
    except Exception as e:
        print(f"  [WARN] Cache write failed: {e}")
    return data


# ============ EXCEL READING ============

class FastExcelReader:
    @staticmethod
    def read_sheet_cached(file_path, sheet_name, cache_name=None, header=0):
        if cache_name is None:
            cache_name = f"{Path(file_path).stem}_{sheet_name}"

        def loader():
            print(f"  [LOCAL EXCEL] Reading {file_path} -> {sheet_name}...")
            reader = ExcelReader(str(file_path))
            return reader.read_sheet(sheet_name=sheet_name, header=header)

        return load_cached(cache_name, loader)


# ============ VARIANT GENERATION ============

def generate_variants_fast(article_df, sku_generator, variant_generator,
                           target_article, target_article_numeric, target_pack,
                           target_brand, target_model, target_brand_short,
                           is_printed=False, printed_articles=None, printed_pack_type="1PC",
                           termination_client=None):
    """Generate SKU variants from merged tracker+attribute data."""

    CUP_COLUMNS    = ["A", "B", "C", "D", "E", "F"]
    STANDARD_SIZES = {"XS", "S", "M", "L", "XL", "XXL", "XXXL", "2XL", "3XL", "4XL", "FS", "FREE SIZE"}

    is_pre_filtered = len(article_df) <= 2 and "_sku_prefix" in article_df.columns
    if "Styles" not in article_df.columns and len(article_df) <= 5:
        is_pre_filtered = True

    filtered_df = article_df.copy() if is_pre_filtered else article_df.copy()

    if "PC" in filtered_df.columns:
        filtered_df["pack_norm"] = filtered_df["PC"].astype(str).str.strip().str.upper()
    elif "pack_norm" not in filtered_df.columns:
        filtered_df["pack_norm"] = target_pack

    all_variants = []

    # ── PRINTED mode ──────────────────────────────────────────────────────────
    if is_printed and printed_articles:
        print(f"\n[PRINTED MODE] Generating variants for {len(printed_articles)} articles...")
        for printed_article in printed_articles:
            brand_short, article_numeric = SkuGenerator.parse_article_string(printed_article)
            matching_rows = []

            if "_article_numeric" in article_df.columns:
                matching_df = article_df[article_df["_article_numeric"] == article_numeric]
                if len(matching_df) > 0:
                    matching_rows = [row for _, row in matching_df.iterrows()]

            if not matching_rows and "Styles" in article_df.columns:
                for idx, row in article_df.iterrows():
                    if article_numeric in str(row.get("Styles", "")):
                        matching_rows.append(row)

            if not matching_rows:
                print(f"    [WARN] No data for article {printed_article}, skipping")
                continue

            row_dict = matching_rows[0].to_dict() if hasattr(matching_rows[0], "to_dict") else matching_rows[0]
            pack  = target_pack if target_pack != "PRINTED" else printed_pack_type
            mrp   = row_dict.get("MRP")
            sp    = row_dict.get("SP")

            for cup in CUP_COLUMNS:
                range_str = str(row_dict.get(cup) or row_dict.get(f"{cup} ") or "").strip()
                if not range_str or "-" not in range_str:
                    continue
                try:
                    start, end = map(int, range_str.split("-"))
                except Exception:
                    continue

                for size_cm in range(start, end + 1, 5):
                    size_inches  = int(round((size_cm / 5) * 2))
                    size_cup_sku = f"{size_inches}{cup}"
                    try:
                        sku_record = sku_generator.generate_from_components(
                            brand_name=target_brand, article_code=printed_article,
                            article_numeric=article_numeric, color_name="PRINTED",
                            pack=pack, size_cup=size_cup_sku, model_name=target_model,
                            brand_short=brand_short
                        )
                        sku_record.update({"mrp": mrp, "sp": sp})
                        variant = variant_generator.generate_from_row(sku_record, is_myntra=True)
                        variant["Color"] = "Printed"
                        variant["Brand Color"] = "Printed"
                        variant["Brand Colour (Remarks)"] = "Printed"
                        variant["Prominent Colour"] = "Multi"
                        for k, val in row_dict.items():
                            if k not in ["seller_sku_id", "vendorSkuCode", "Color", "Brand Color"]:
                                if k not in variant or not variant[k]:
                                    variant[k] = val
                        if mrp: variant["MRP"] = mrp
                        if sp:  variant["SP"]  = sp
                        variant["_cup"]  = cup
                        variant["_size"] = size_cup_sku
                        all_variants.append(variant)
                    except Exception as e:
                        print(f"    [WARN] Error for {printed_article} {size_cup_sku}: {e}")
        print(f"  [OK] Generated {len(all_variants)} variants (PRINTED mode)")

    # ── Normal mode ───────────────────────────────────────────────────────────
    if not (is_printed and printed_articles):
        for idx, row in filtered_df.iterrows():
            ensure_not_terminated(termination_client, f"processing row {idx}")

            excel_article = row.get("Styles")
            if not excel_article or str(excel_article).strip() == "":
                brand_col_value = row.get(target_brand, "")
                if brand_col_value and str(brand_col_value).strip():
                    parts = str(brand_col_value).split()
                    if parts:
                        excel_article = f"{target_brand_short}-{parts[-1]}"
                if not excel_article or str(excel_article).strip() == "":
                    excel_article = target_article

            color = row.get("Colors") or row.get("COLOR") or "UNKNOWN"
            article_numeric = target_article_numeric
            article = target_article
            pack    = row["pack_norm"]
            mrp     = row.get("MRP")
            sp      = row.get("SP")

            image_urls = {
                "Main Image URL":    row.get("Main Image URL", ""),
                "Other Image URL 1": row.get("Other Image URL 1", ""),
                "Other Image URL 2": row.get("Other Image URL 2", ""),
                "Other Image URL 3": row.get("Other Image URL 3", ""),
                "Other Image URL 4": row.get("Other Image URL 4", ""),
            }

            skip_columns = {
                "Styles", "COLOR", "Colors", "PC", "MRP", "SP",
                "A", "B", "C", "D", "E", "F", "pack_norm", "_sku_prefix",
                "Fabric 2", "Fabric 3", "Sports Bra Support", "Technology", "Sport",
                "Multipack Set", "Number of Items", "Package Contains", "Net Quantity",
                "Brand Colour (Remarks)", "Prominent Colour",
                "productDisplayName", "Net Quantity Unit"
            }

            for cup in CUP_COLUMNS:
                range_str = str(row.get(cup) or row.get(f"{cup} ") or "").strip()

                # Standard size (S, M, L, etc.)
                if range_str.upper() in STANDARD_SIZES:
                    size_cup_sku = range_str.upper()
                    sku_record = sku_generator.generate_from_components(
                        brand_name=target_brand, article_code=article,
                        article_numeric=article_numeric, color_name=color,
                        pack=pack, size_cup=size_cup_sku, model_name=target_model,
                        brand_short=target_brand_short
                    )
                    sku_record.update({"mrp": mrp, "sp": sp,
                                       "vendorArticleName": row.get("vendorArticleName", ""),
                                       **image_urls})
                    variant_row = variant_generator.generate_from_row(sku_record)
                    for col_name, col_value in row.items():
                        if col_name in skip_columns:
                            continue
                        existing = variant_row.get(col_name, "")
                        if col_name not in variant_row or str(existing).strip() == "":
                            variant_row[col_name] = col_value
                    variant_row["_cup"]  = cup
                    variant_row["_size"] = size_cup_sku
                    all_variants.append(variant_row)
                    continue

                if not range_str or "-" not in range_str:
                    continue

                try:
                    start, end = map(int, range_str.split("-"))
                except Exception:
                    continue

                for size_cm in range(start, end + 1, 5):
                    size_inches  = int(round((size_cm / 5) * 2))
                    size_cup_sku = f"{size_inches}{cup}"
                    sku_record = sku_generator.generate_from_components(
                        brand_name=target_brand, article_code=article,
                        article_numeric=article_numeric, color_name=color,
                        pack=pack, size_cup=size_cup_sku, model_name=target_model,
                        brand_short=target_brand_short
                    )
                    sku_record.update({"mrp": mrp, "sp": sp,
                                       "vendorArticleName": row.get("vendorArticleName", ""),
                                       **image_urls})
                    variant_row = variant_generator.generate_from_row(sku_record)
                    for col_name, col_value in row.items():
                        if col_name in skip_columns:
                            continue
                        existing = variant_row.get(col_name, "")
                        if col_name not in variant_row or str(existing).strip() == "":
                            variant_row[col_name] = col_value
                    variant_row["_cup"]  = cup
                    variant_row["_size"] = size_cup_sku
                    all_variants.append(variant_row)

    print(f"  [OK] Generated {len(all_variants)} variants")
    return all_variants


# ============ FILE WRITING (no dropdown preservation for Flipkart) ============

def write_excel_fast(template_path, df, output_path, sheet_name, header_row, start_row):
    wb_template = load_workbook(template_path)
    if sheet_name not in wb_template.sheetnames:
        raise FileNotFoundError(f"Sheet '{sheet_name}' not found in template")
    ws = wb_template[sheet_name]

    header_row = int(header_row)
    start_row  = int(start_row)

    # Snapshot formatting before any modification
    saved_col_widths        = {k: v.width  for k, v in ws.column_dimensions.items()}
    saved_row_heights       = {k: v.height for k, v in ws.row_dimensions.items()}
    saved_default_col_width = ws.sheet_format.defaultColWidth
    saved_default_row_height = ws.sheet_format.defaultRowHeight

    # Clear existing data rows only
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

    # Build header -> column index map
    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        hval = ws.cell(row=header_row, column=col_idx).value
        if hval:
            header_to_col[str(hval).strip()] = col_idx

    left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)

    NUMERIC_COLS = {
        "MRP (INR)", "Your selling price (INR)", "Group ID",
        "Overbust Range ( Inches )", "Underbust Range ( Inches )",
    }

    records = df.to_dict(orient="records")
    for r_idx, row in enumerate(records, start=start_row):
        for col_name, value in row.items():
            clean = str(col_name).strip() if col_name else None
            col_idx = header_to_col.get(clean)
            if not col_idx:
                continue
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            cell.alignment = left_align
            if col_name in NUMERIC_COLS and value:
                try:
                    num = float(str(value).replace(",", "").strip()) if isinstance(value, str) else float(value)
                    cell.value = int(num) if num == int(num) else num
                    cell.number_format = "#,##0"
                except Exception:
                    pass

    # Restore template formatting exactly
    if saved_default_col_width:
        ws.sheet_format.defaultColWidth = saved_default_col_width
    if saved_default_row_height:
        ws.sheet_format.defaultRowHeight = saved_default_row_height
    for col_letter, width in saved_col_widths.items():
        ws.column_dimensions[col_letter].width = width
    for row_num, height in saved_row_heights.items():
        ws.row_dimensions[row_num].height = height

    wb_template.save(output_path)
    print(f"  [OK] Saved: {Path(output_path).name} ({len(df)} rows)")


# ============ COLUMN MAPPER ============

class TemplateColumnMapper:
    def __init__(self, template_df: pd.DataFrame):
        self.template_columns = list(template_df.columns)
        if not self.template_columns:
            raise ValueError("Template has no columns")

    def map_rows(self, rows: list) -> pd.DataFrame:
        mapped_rows = []
        for row in rows:
            flat = {}
            for k, v in row.items():
                if k is None:
                    continue
                key = str(k).strip()
                flat[key] = v.get("value", "") if isinstance(v, dict) else v
            mapped_rows.append({col: flat.get(str(col).strip(), "") for col in self.template_columns})
        return pd.DataFrame(mapped_rows, columns=self.template_columns)


# ============ MAIN ============

def main():
    print("\n" + "=" * 70)
    print("[START] FLIPKART Listing Generator")
    print("[MODE]  Multi-workbook: Flipkart Tracker + Brand Attributes")
    print("=" * 70)

    overall_start = time.time()

    # Initialize template
    print("\n[0/8] Initializing Flipkart template...")
    initialize_template()
    if FLIPKART_TEMPLATE and FLIPKART_TEMPLATE_SHEET:
        print(f"  [OK] Template ready: {FLIPKART_TEMPLATE.name} ({FLIPKART_TEMPLATE_SHEET})")
    else:
        print("  [ERROR] Could not initialize Flipkart template")
        sys.exit(1)

    # Load config
    print("\n[1/8] Loading configuration...")
    start = time.time()
    with open(CONFIG_DIR / "run_config.yaml", "r", encoding="utf-8") as f:
        run_config = yaml.safe_load(f)

    TARGET_ARTICLE     = run_config["article"]
    TARGET_PACK        = run_config["pack"].upper()
    PRINTED_PACK_TYPE  = run_config.get("printed_pack_type", "1PC").upper()
    TARGET_BRAND       = run_config.get("brand", "INVISI-SOFT")
    TARGET_MODEL       = run_config.get("model", "MAGDHA")

    # Flipkart brand key = "fk_" + normalised brand name
    BRAND_NORMALIZED = "fk_" + TARGET_BRAND.lower().replace("-", "")

    # PRINTED range?
    IS_PRINTED_RANGE = False
    PRINTED_ARTICLES = []
    if TARGET_PACK == "PRINTED" and " to " in TARGET_ARTICLE:
        IS_PRINTED_RANGE = True
        try:
            from_article, to_article = TARGET_ARTICLE.split(" to ")
            from_article, to_article = from_article.strip(), to_article.strip()
            from_parts = from_article.rsplit("-", 1)
            to_parts   = to_article.rsplit("-", 1)
            brand_prefix = from_parts[0]
            for num in range(int(from_parts[-1]), int(to_parts[-1]) + 1):
                PRINTED_ARTICLES.append(f"{brand_prefix}-{num}")
            print(f"  [PRINTED MODE] Range: {PRINTED_ARTICLES}")
            TARGET_ARTICLE = PRINTED_ARTICLES[0]
        except Exception as e:
            print(f"  [ERROR] Failed to parse article range: {e}")
            sys.exit(1)

    from engine.sku_generator import SkuGenerator
    try:
        TARGET_BRAND_SHORT, TARGET_ARTICLE_NUMERIC = SkuGenerator.parse_article_string(TARGET_ARTICLE)
        print(f"  [OK] Article: brand_short='{TARGET_BRAND_SHORT}', numeric='{TARGET_ARTICLE_NUMERIC}'")
    except Exception as e:
        print(f"  [ERROR] Failed to parse article '{TARGET_ARTICLE}': {e}")
        sys.exit(1)

    generation_status = None
    if STATUS_REPORTING_AVAILABLE:
        generation_status = GenerationStatus(TARGET_ARTICLE, TARGET_BRAND, OUTPUT_DIR)

    now   = datetime.now()
    MONTH = now.strftime("%m")
    YEAR  = now.strftime("%y")

    print(f"  [OK] Article: {TARGET_ARTICLE}, Pack: {TARGET_PACK}, Brand: {TARGET_BRAND}, Model: {TARGET_MODEL}")
    print(f"  [OK] SKU Date: {MONTH}{YEAR}  ({time.time()-start:.2f}s)")

    termination_client = initialize_device_client()
    ensure_not_terminated(termination_client, "startup")

    # Rule engine
    print("\n[2/8] Loading rule engine...")
    start = time.time()
    rule_engine = RuleEngine(str(CONFIG_DIR), article_master={})
    print(f"  [OK] Brand prefix: {TARGET_BRAND_SHORT} ({time.time()-start:.2f}s)")

    # Load Flipkart template (cached, keyed with header row to avoid stale reads)
    print("\n[3/8] Loading Flipkart template...")
    start = time.time()
    template_cache_key = f"flipkart_template_hdr{FLIPKART_HEADER_ROW}"
    template_df = FastExcelReader.read_sheet_cached(
        FLIPKART_TEMPLATE, FLIPKART_TEMPLATE_SHEET,
        template_cache_key,
        header=FLIPKART_HEADER_ROW - 1   # pandas 0-indexed
    )
    print(f"  [OK] Done ({time.time()-start:.2f}s)")

    # Generators
    print("\n[4/8] Initializing generators...")
    start = time.time()
    sku_generator     = SkuGenerator(rule_engine=rule_engine, month=MONTH, year=YEAR)
    variant_generator = VariantGenerator(rule_engine=rule_engine)
    column_mapper     = TemplateColumnMapper(template_df)
    print(f"  [OK] Done ({time.time()-start:.2f}s)")

    # Read from Flipkart tracker + brand attributes
    print(f"\n[5/8] Reading from Flipkart tracker and {TARGET_BRAND} workbook...")
    start = time.time()

    if not (MULTI_WORKBOOK_AVAILABLE and BRAND_NORMALIZED in ALL_VALID_BRANDS):
        print(f"  [ERROR] Brand '{BRAND_NORMALIZED}' not in supported list: {ALL_VALID_BRANDS}")
        sys.exit(1)

    try:
        config_mgr   = BrandConfigManager()
        reader       = MultiWorkbookReader(config_mgr)
        brand_config = config_mgr.get_brand_config(BRAND_NORMALIZED)

        print(f"  [DEBUG] SKU sheet:  {brand_config['sku_source']['sheet_name']}")
        print(f"  [DEBUG] Attr sheet: {brand_config['attribute_source']['sheet_name']}")

        if IS_PRINTED_RANGE and PRINTED_ARTICLES:
            all_article_data = []
            for article in PRINTED_ARTICLES:
                try:
                    article_df, metadata = reader.get_merged_data(BRAND_NORMALIZED, article)
                    if len(article_df) > 0:
                        _, art_num = SkuGenerator.parse_article_string(article)
                        article_df["_article_numeric"] = art_num
                        all_article_data.append(article_df)
                        print(f"    [OK] {article}: {len(article_df)} rows")
                    else:
                        print(f"    [WARN] No data for {article}")
                except Exception as e:
                    print(f"    [WARN] Error loading {article}: {e}")
            if not all_article_data:
                print("  [ERROR] No data for any article in range")
                sys.exit(1)
            merged_df = pd.concat(all_article_data, ignore_index=True)
        else:
            merged_df, metadata = reader.get_merged_data(BRAND_NORMALIZED, TARGET_ARTICLE)
            if not IS_PRINTED_RANGE:
                if metadata.get("sku_rows", 0) == 0:
                    print(f"\n{'='*70}\n[GENERATION SKIPPED] SKU not found in Flipkart Tracker\n{'='*70}")
                    sys.exit(0)
                if metadata.get("attr_rows", 0) == 0:
                    print(f"\n{'='*70}\n[GENERATION SKIPPED] Attributes not found in brand workbook\n{'='*70}")
                    sys.exit(0)

        multi_workbook_data = merged_df.copy()
        print(f"  [OK] {len(merged_df)} rows loaded ({time.time()-start:.2f}s)")

    except Exception as e:
        print(f"  [WARN] Error reading data: {e}")
        traceback.print_exc()
        sys.exit(1)

    # Generate variants
    print("\n[6/8] Generating variants...")
    start = time.time()
    all_variants = generate_variants_fast(
        multi_workbook_data, sku_generator, variant_generator,
        TARGET_ARTICLE, TARGET_ARTICLE_NUMERIC, TARGET_PACK,
        TARGET_BRAND, TARGET_MODEL, TARGET_BRAND_SHORT,
        is_printed=IS_PRINTED_RANGE, printed_articles=PRINTED_ARTICLES,
        printed_pack_type=PRINTED_PACK_TYPE,
        termination_client=termination_client
    )
    print(f"  [OK] Done ({time.time()-start:.2f}s)")

    # Assign styleGroupId
    if IS_PRINTED_RANGE and PRINTED_ARTICLES:
        print("\n[6.5/8] Assigning styleGroupId (PRINTED mode — by article)...")
        article_map, next_id = {}, 1
        for v in all_variants:
            art = v.get("article_numeric", "UNKNOWN")
            if art not in article_map:
                article_map[art] = next_id
                next_id += 1
            v["styleGroupId"] = article_map[art]
    else:
        print("\n[6.5/8] Assigning styleGroupId (by color)...")
        color_map, next_id = {}, 1
        for v in all_variants:
            color = v.get("Color", "UNKNOWN")
            if color not in color_map:
                color_map[color] = next_id
                next_id += 1
            v["styleGroupId"] = color_map[color]

    # ── Flipkart column remapping ──────────────────────────────────────────────
    print("\n[6.7/8] Remapping columns to Flipkart template format...")
    pack_num = "2" if "2" in str(TARGET_PACK) else "1"
    for v in all_variants:
        v["Seller SKU ID"]              = v.get("vendorSkuCode", "")
        v["Group ID"]                   = str(v.get("styleGroupId", ""))
        v["MRP (INR)"]                  = str(v.get("MRP", v.get("mrp", ""))).replace(",", "").strip()
        v["Your selling price (INR)"]   = str(v.get("SP", v.get("sp", ""))).replace(",", "").strip()
        v["Color"]                      = v.get("Color") or v.get("Colors", "")
        v["Brand Color"]                = v.get("Brand Colour (Remarks)", v.get("Brand Color", ""))
        v["Size"]                       = v.get("_size", "")
        v["Size - Measuring Unit"]      = "Number"
        v["Style Code"]                 = TARGET_ARTICLE_NUMERIC
        v["Pack of"]                    = pack_num
        v["Fabric Care"]                = v.get("materialCareDescription", v.get("Fabric Care", ""))
        v["Description"]                = v.get("Product Details", v.get("Description", ""))
        v["Manufacturer Details"]       = v.get("Manufacturer Name and Address with Pincode", v.get("Manufacturer Details", ""))
        v["Packer Details"]             = v.get("Packer Name and Address with Pincode",       v.get("Packer Details", ""))
        v["Importer Details"]           = v.get("Importer Name and Address with Pincode",     v.get("Importer Details", ""))
    print(f"  [OK] Remapped {len(all_variants)} variants")

    # Map columns and write output
    ensure_not_terminated(termination_client, "before mapping columns")
    print("\n[7/8] Mapping columns and writing output...")
    start = time.time()

    # Group variants by article
    article_groups = defaultdict(list)
    if IS_PRINTED_RANGE and PRINTED_ARTICLES:
        group_name = (f"{PRINTED_ARTICLES[0]}_to_{PRINTED_ARTICLES[-1]}"
                      if len(PRINTED_ARTICLES) > 1 else PRINTED_ARTICLES[0])
        for v in all_variants:
            article_groups[group_name].append(v)
    else:
        for v in all_variants:
            seller_sku      = v.get("vendorSkuCode", "")
            article_numeric = v.get("article_numeric", "")
            if seller_sku and article_numeric:
                sku_no_model = seller_sku.split("_")[0] if "_" in seller_sku else seller_sku
                parts        = sku_no_model.split("-")
                article_code = None
                for i, part in enumerate(parts):
                    if article_numeric in part:
                        article_code = "-".join(parts[:i+1])
                        break
                article_code = article_code or parts[0]
            else:
                article_code = "UNKNOWN"
            article_groups[article_code].append(v)

    # Output folder
    output_folder = Path(OUTPUT_DIR)
    try:
        cfg_mgr      = BrandConfigManager()
        brand_cfg    = cfg_mgr.get_brand_config(BRAND_NORMALIZED)
        brand_folder = Path(brand_cfg.get("output_folder", f"data/output/flipkart/{BRAND_NORMALIZED}/"))
        output_folder = BASE_DIR / brand_folder
    except Exception as e:
        print(f"  [WARN] Could not determine brand folder: {e}")
    output_folder.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    IMPORTANT_COLS = ["Seller SKU ID", "Group ID", "MRP (INR)", "Your selling price (INR)", "Color", "Size", "Style Code"]

    for article, variants in article_groups.items():
        total = len(variants)
        print(f"\n  Processing {article} ({total} SKUs)...")
        mapped_df = column_mapper.map_rows(variants)

        for col in IMPORTANT_COLS:
            if col in mapped_df.columns:
                filled = mapped_df[col].astype(str).str.strip().replace("nan", "").ne("").sum()
                if filled == 0:
                    print(f"  [WARN EMPTY_COL] '{col}' — no data in all {total} rows, check attribute sheet")

        out_name = output_folder / f"Flipkart_Sku_Ready_{article}_{TARGET_PACK}_{ts}.xlsx"
        print(f"    -> {out_name.name}")
        write_excel_fast(
            str(FLIPKART_TEMPLATE), mapped_df, str(out_name),
            sheet_name=FLIPKART_TEMPLATE_SHEET,
            header_row=FLIPKART_HEADER_ROW,
            start_row=FLIPKART_START_ROW
        )

    print(f"  [OK] Done ({time.time()-start:.2f}s)")

    # Summary
    file_count = len(list(output_folder.glob("Flipkart_Sku_Ready_*.xlsx")))
    total_time = time.time() - overall_start
    print("\n" + "=" * 70)
    print(f"[SUCCESS] ALL COMPLETE in {total_time:.2f}s ({total_time/60:.1f}min)")
    print(f"[OUTPUT]  {file_count} file(s) saved to: {output_folder}")
    print("=" * 70)

    if generation_status:
        generation_status.success(file_count)
        generation_status.save_status()

    if ENABLE_CACHE:
        print(f"\n[TIP] Next run will be faster (cache: {CACHE_DIR})")


if __name__ == "__main__":
    try:
        main()
    except ValidationError as ve:
        print(f"\n[FAIL] VALIDATION ERROR: {ve}")
        sys.exit(1)
    except Exception as e:
        print(f"\n[FAIL] ERROR: {e}")
        traceback.print_exc()
        sys.exit(1)
